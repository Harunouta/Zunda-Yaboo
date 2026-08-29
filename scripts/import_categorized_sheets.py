"""Import shaped / conflict / policy sheets into japan events and policy catalog.

Skips インポート済. Imports 争い and 政策 even without 整形済.
statute_* rows go to data/events/policies/catalog.yaml (not monthly shocks).
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ENV_XLSX = "ZUNDA_EVENTS_XLSX"
DEFAULT_XLSX_NAMES = (
  "categorized_events.xlsx",
)
JAPAN = ROOT / "data" / "events" / "japan"
POLICIES = ROOT / "data" / "events" / "policies"
LOGS = ROOT / "logs"

SHAPED_MARK = "整形済"
SHAPED_ALIASES = ("整形済", "整形願", "追記して")
SKIP_MARK = "インポート済"
SKIP_SHEET_TOKENS = ("水害", "飢饉", "農作物")
FORCE_IMPORT_TOKENS = ("争い", "政策")
SIM_START = "1603-01"

WAR_NOTE_MARKERS = ("戦争", "陣", "乱", "討伐", "合戦")
IKKI_NOTE_MARKERS = ("一揆", "騒動", "蜂起")

# Major volcanoes: apply harvest / ash shocks. Others get prompts only.
MAJOR_NOTE_MARKERS = (
  "宝永",
  "島原大変",
  "大正大噴火",
  "浅間",
  "磐梯",
  "御嶽",
  "三宅",
  "有珠",
  "新燃",
  "雲仙",
  "桜島",
)
MAJOR_ID_MARKERS = (
  "hoei",
  "unzen",
  "sakurajima",
  "asama",
  "bandai",
  "ontake",
  "miyake",
  "usu_",
  "sinmoe",
  "shinmoe",
)

EPIDEMIC_SEVERITY = {
  "bunsei_cholera": 0.35,
  "ansei_cholera": 0.55,
  "russian_flu": 0.22,
  "spanish_flu": 0.6,
  "asian_flu": 0.28,
  "hong_kong_flu": 0.25,
  "swine_flu_2009": 0.12,
  "covid_19_start": 0.2,
  "covid_19_delta": 0.22,
  "covid_19_omicron": 0.15,
}


def normalizeYearMonth(raw: object) -> str:
  if isinstance(raw, datetime):
    return f"{raw.year:04d}-{raw.month:02d}"
  text = str(raw).strip()
  if "T" in text:
    text = text.split("T", 1)[0]
  text = text.replace("_", "-")
  yearMatch = re.search(r"(\d{3,4})", text)
  monthMatch = re.search(r"年\s*(\d{1,2})", text)
  if yearMatch and ("年" in text):
    year = int(yearMatch.group(1))
    month = int(monthMatch.group(1)) if monthMatch else 1
    return f"{year:04d}-{month:02d}"
  parts = text.split("-")
  if len(parts) >= 2 and parts[0].isdigit() and parts[1][:2].isdigit():
    return f"{int(parts[0]):04d}-{int(parts[1][:2]):02d}"
  raise ValueError(f"bad yearMonth: {raw!r}")


def normalizeEventId(raw: str) -> str:
  text = str(raw).strip()
  text = text.replace("-", "_").replace(" ", "_")
  text = re.sub(r"[^0-9A-Za-z_]", "_", text)
  text = re.sub(r"_+", "_", text).strip("_")
  return text.lower()


def decodeSource(url: str | None) -> str:
  if not url:
    return ""
  return unquote(str(url))


def optionalFloat(value: object) -> float | None:
  if value is None or value == "":
    return None
  try:
    return float(value)
  except (TypeError, ValueError):
    return None


def resolveXlsx(cliPath: str | None) -> Path:
  if cliPath:
    path = Path(cliPath)
    if not path.is_file():
      raise SystemExit(f"missing xlsx: {path}")
    return path
  envPath = os.getenv(ENV_XLSX)
  if envPath:
    path = Path(envPath)
    if not path.is_file():
      raise SystemExit(f"missing xlsx ({ENV_XLSX}): {path}")
    return path
  downloads = Path.home() / "Downloads"
  tried: list[Path] = []
  for name in DEFAULT_XLSX_NAMES:
    path = downloads / name
    tried.append(path)
    if path.is_file():
      return path
  hint = " or ".join(str(path) for path in tried)
  raise SystemExit(f"missing xlsx: pass --xlsx PATH or set {ENV_XLSX} (tried {hint})")


def shouldSkipSheet(sheetName: str) -> bool:
  if SKIP_MARK in sheetName:
    return True
  if any(token in sheetName for token in FORCE_IMPORT_TOKENS):
    return False
  if not any(mark in sheetName for mark in SHAPED_ALIASES):
    return True
  return any(token in sheetName for token in SKIP_SHEET_TOKENS)


def collectLooseUrls(rows: list[tuple]) -> list[str]:
  """URLs on rows with no yearMonth (sheet-footer bibliography)."""
  urls: list[str] = []
  for row in rows[1:]:
    if not row:
      continue
    hasYearMonth = row[0] is not None and str(row[0]).strip() != ""
    if hasYearMonth:
      continue
    for cell in row:
      if cell is None:
        continue
      text = str(cell).strip()
      if text.startswith("http") and text not in urls:
        urls.append(text)
  return urls


def dumpShapedSheets(xlsxPath: Path) -> tuple[dict[str, list[dict]], list[str]]:
  workbook = load_workbook(xlsxPath, read_only=True, data_only=True)
  out: dict[str, list[dict]] = {}
  bibliography: list[str] = []
  for sheetName in workbook.sheetnames:
    if shouldSkipSheet(sheetName):
      continue
    worksheet = workbook[sheetName]
    rows = list(worksheet.iter_rows(values_only=True))
    bibliography.extend(collectLooseUrls(rows))
    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    items: list[dict] = []
    for row in rows[1:]:
      if not row or row[0] is None or str(row[0]).strip() == "":
        continue
      entry: dict = {}
      extraUrls: list[str] = []
      for key, value in zip(header, row):
        if key in ("disasterOverride", "populationShock", "cropLossExtra"):
          entry[key] = optionalFloat(value)
          continue
        if value is None or value == "":
          if key:
            entry[key] = None
          continue
        text = value.isoformat() if isinstance(value, datetime) else str(value).strip()
        if not key or key in ("None", "nan"):
          if text.startswith("http"):
            extraUrls.append(text)
          continue
        entry[key] = text
      if not entry.get("sourceUrl") and extraUrls:
        entry["sourceUrl"] = extraUrls[0]
      if entry.get("yearMonth") and entry.get("eventId"):
        items.append(entry)
    out[sheetName] = items
  # Dedupe bibliography preserving order
  seenBib: set[str] = set()
  uniqueBib: list[str] = []
  for url in bibliography:
    if url in seenBib:
      continue
    seenBib.add(url)
    uniqueBib.append(url)
  return out, uniqueBib


def seriesIndex(eventId: str) -> int | None:
  match = re.search(r"_(\d+)$", eventId)
  if not match:
    return None
  return int(match.group(1))


def volcanoTier(eventId: str, notes: str, yearMonth: str) -> str:
  blobId = eventId.lower()
  year = int(yearMonth[:4])
  if any(marker in notes for marker in MAJOR_NOTE_MARKERS) or any(
    marker in blobId for marker in MAJOR_ID_MARKERS
  ):
    if "浅間" in notes or "asama" in blobId:
      if year == 1783:
        return "major"
      return "moderate"
    if "桜島" in notes or "sakurajima" in blobId:
      if "taisho" in blobId or "大正" in notes or "anei" in blobId or "showa" in blobId:
        return "major"
      return "moderate"
    if "usu" in blobId or "有珠" in notes:
      return "moderate"
    return "major"
  return "minor"


def fillVolcano(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  eventId = row["eventId"]
  yearMonth = row["yearMonth"]
  source = decodeSource(row.get("sourceUrl"))
  index = seriesIndex(eventId)
  tier = volcanoTier(eventId, notes, yearMonth)
  display = re.sub(r"_\d+$", "", notes)

  if eventId.startswith("hoei_fuji") or "宝永" in notes:
    leader = "富士山が噴火した。江戸にも灰が降り、田畑が埋まる。救恤と年貢減免を急げ。"
    opinion = "昼なのに空が暗い。灰が積もって息が苦しい。"
    coeffs = (0.4, None, 0.2)
  elif eventId == "unzen_tsunami" or "島原大変" in notes:
    leader = "肥前で山が崩れ、対岸まで津波が襲った。甚大な被害への救済を行え。"
    opinion = "山が海に落ち、波が村をさらった。"
    coeffs = (0.4, 0.012, 0.08)
  elif "sakurajima_taisho" in eventId or "大正大噴火" in notes:
    leader = "桜島が大噴火し、島が陸続きになった。避難民の受け入れと降灰対策を行え。"
    opinion = "山が火を噴き、灰で昼も暗い。"
    coeffs = (0.45, None, 0.1)
  elif "浅間" in notes and yearMonth.startswith("1783"):
    if index is None or index <= 4:
      leader = "浅間山が激しく噴火している。灰が田を覆う。年貢と備蓄をどうする？"
      opinion = "空が昼でも暗い。作物が死ぬと噂が広がる。"
      coeffs = (0.4, 0.004, 0.18)
    else:
      leader = "浅間の噴火が続く。灰害の余波で作柄が崩れている。救済を続けよ。"
      opinion = "まだ灰が降る。米が足りないと噂だ。"
      coeffs = (0.5, 0.002, 0.1)
  elif "磐梯" in notes or "bandai" in eventId:
    leader = f"{display}。山体崩壊の報がある。被災地の救援と交通の確保を急げ。"
    opinion = "山が吹き飛んだ。川が堰き止められると噂だ。"
    coeffs = (0.48, 0.006, 0.08)
  elif "御嶽" in notes or "ontake" in eventId:
    leader = f"{display}。登山客と山麓の被災への救援を急げ。"
    opinion = "山が突然噴いた。逃げ遅れた人がいる。"
    coeffs = (0.62, 0.002, 0.03)
  elif tier == "major":
    leader = f"{display}の噴火だ。降灰・避難・農地被害への対応を命じよ。"
    opinion = "山が鳴っている。灰が降ると作が危ない。"
    coeffs = (0.5, 0.002, 0.08)
  elif tier == "moderate":
    leader = f"{display}が活動している。降灰と交通、作柄への影響を監視せよ。"
    opinion = "山から煙が立つ。灰が洗濯物に付く。"
    coeffs = (0.7, None, 0.03)
  else:
    leader = f"{display}の小規模な噴火・活動だ。観測と周辺の安全を確認せよ。"
    opinion = "山が少し騒がしい。普段と違う匂いがする。"
    coeffs = (None, None, None)

  if not row.get("promptForLeader"):
    row["promptForLeader"] = leader
  if not row.get("promptForOpinionLeader"):
    row["promptForOpinionLeader"] = opinion
  if row.get("disasterOverride") is None and coeffs[0] is not None:
    row["disasterOverride"] = coeffs[0]
  if row.get("populationShock") is None and coeffs[1] is not None:
    row["populationShock"] = coeffs[1]
  if row.get("cropLossExtra") is None and coeffs[2] is not None:
    row["cropLossExtra"] = coeffs[2]
  row["_sourceDecoded"] = source
  row["_kind"] = "volcano"
  return row


def fillInfra(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  if not row.get("promptForLeader"):
    row["promptForLeader"] = f"{notes}。流通・情報・財政への波及を見定め、秩序を保て。"
  if not row.get("promptForOpinionLeader"):
    row["promptForOpinionLeader"] = f"{notes}、と市で噂になっている。暮らしが変わるのか。"
  # Informational: do not invent harvest shocks (coeffs may change later).
  row["_sourceDecoded"] = decodeSource(row.get("sourceUrl"))
  row["_kind"] = "infra"
  return row


def fillEpidemic(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  eventId = row["eventId"]
  severity = EPIDEMIC_SEVERITY.get(eventId, 0.2)
  if eventId in ("bunsei_cholera", "ansei_cholera"):
    leader = f"{notes}。感染を抑え、埋葬と救恤、米の流通を途切れさせるな。"
    opinion = "コロリが来る。水も人も疑え、と囁く。"
  elif "flu" in eventId or "かぜ" in notes or "インフル" in notes:
    leader = f"{notes}。病欠と物流停滞に備え、救済と噂の抑制を。"
    opinion = "熱が出る人が増えた。市が寂しい。"
  else:
    leader = f"{notes}。感染拡大を抑え、民の不安と経済の停滞を両立して扱え。"
    opinion = "病の噂が先に走る。外に出るのが怖い。"
  if not row.get("promptForLeader"):
    row["promptForLeader"] = leader
  if not row.get("promptForOpinionLeader"):
    row["promptForOpinionLeader"] = opinion
  row["epidemicSeverity"] = severity
  row["_sourceDecoded"] = decodeSource(row.get("sourceUrl"))
  row["_kind"] = "epidemic"
  return row


def fillEarthquake(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  display = re.sub(r"_\d+$", "", notes)
  swarm = "群発" in notes
  tsunami = "津波" in notes
  if not row.get("promptForLeader"):
    if tsunami:
      row["promptForLeader"] = f"{display}。津波と沿岸被害の報がある。救恤と復旧を急げ。"
    elif swarm:
      row["promptForLeader"] = f"{display}。揺れが続いている。建物と人心の安定を図れ。"
    else:
      row["promptForLeader"] = f"{display}。被害の把握と救援、交通の確保を急げ。"
  if not row.get("promptForOpinionLeader"):
    if tsunami:
      row["promptForOpinionLeader"] = "海が引いたあと波が来た、と港で叫ぶ。"
    elif swarm:
      row["promptForOpinionLeader"] = "また揺れた。蔵の甕が鳴る。"
    else:
      row["promptForOpinionLeader"] = "大地が揺れた。家が傾いたと噂だ。"
  if swarm:
    pass
  elif tsunami:
    if row.get("disasterOverride") is None:
      row["disasterOverride"] = 0.55
    if row.get("cropLossExtra") is None:
      row["cropLossExtra"] = 0.04
    if row.get("populationShock") is None:
      row["populationShock"] = 0.003
  row["_sourceDecoded"] = decodeSource(row.get("sourceUrl"))
  row["_kind"] = "earthquake"
  return row


def fillFire(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  display = re.sub(r"[_（].*$", "", notes)
  bigFire = "大火" in notes or "焼失" in notes
  forestFire = "山林" in notes
  if not row.get("promptForLeader"):
    if bigFire:
      row["promptForLeader"] = f"{display}。市街が広く焼けた。消火の後の復興と治安を急げ。"
    elif forestFire:
      row["promptForLeader"] = f"{display}。山火事の拡大を抑え、集落の避難を急げ。"
    else:
      row["promptForLeader"] = f"{display}。火災の被害把握と救援を行え。"
  if not row.get("promptForOpinionLeader"):
    if bigFire:
      row["promptForOpinionLeader"] = "町が燃えている。風で火の粉が飛ぶ。"
    elif forestFire:
      row["promptForOpinionLeader"] = "山が赤い。煙で空がかすむ。"
    else:
      row["promptForOpinionLeader"] = "火事だ、と人が走る。"
  if bigFire:
    if row.get("disasterOverride") is None:
      row["disasterOverride"] = 0.5
    if row.get("cropLossExtra") is None:
      row["cropLossExtra"] = 0.02
  elif forestFire:
    if row.get("disasterOverride") is None:
      row["disasterOverride"] = 0.65
    if row.get("cropLossExtra") is None:
      row["cropLossExtra"] = 0.04
  row["_sourceDecoded"] = decodeSource(row.get("sourceUrl"))
  row["_kind"] = "fire"
  return row


def fillConflict(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  display = re.sub(r"_\d+$", "", notes)
  eventId = row["eventId"]
  blob = f"{eventId} {notes}"
  isWar = any(marker in blob for marker in WAR_NOTE_MARKERS) or "_war" in eventId
  isIkki = any(marker in blob for marker in IKKI_NOTE_MARKERS)
  index = seriesIndex(eventId)
  continuation = index is not None and index >= 2
  if not row.get("promptForLeader"):
    if isIkki:
      row["promptForLeader"] = f"{display}。一揆・騒動の報がある。年貢と兵の扱いを誤るな。"
    elif isWar:
      row["promptForLeader"] = f"{display}。戦が広がっている。兵糧と治安、民の保護を急げ。"
    else:
      row["promptForLeader"] = f"{display}。争いの報がある。鎮圧と救済のバランスを取れ。"
  if not row.get("promptForOpinionLeader"):
    if isIkki:
      row["promptForOpinionLeader"] = "百姓が起った、と村から噂が来る。"
    else:
      row["promptForOpinionLeader"] = "戦だ。兵が通ると蔵が空く。"
  if row.get("disasterOverride") is None and not continuation:
    if "shimabara" in eventId or "島原" in notes:
      row["disasterOverride"] = 0.6
    elif isWar and not isIkki:
      row["disasterOverride"] = 0.72
  if row.get("populationShock") is None:
    if continuation:
      row["populationShock"] = 0.001
    elif "shimabara" in eventId or "島原" in notes:
      row["populationShock"] = 0.008
    elif isWar and not isIkki:
      row["populationShock"] = 0.005
    elif isIkki:
      row["populationShock"] = 0.0015
  if row.get("cropLossExtra") is None and (isWar or isIkki) and not continuation:
    row["cropLossExtra"] = 0.04 if isWar and not isIkki else 0.02
  if row.get("socialUnrest") is None:
    row["socialUnrest"] = 0.04 if continuation else (0.14 if isWar and not isIkki else 0.08)
  if row.get("laborDrain") is None:
    row["laborDrain"] = 0.015 if continuation else 0.03
  if row.get("infraDamage") is None and isWar and not continuation:
    row["infraDamage"] = 0.04
  row["scope"] = row.get("scope") or "japan"
  row["_sourceDecoded"] = decodeSource(row.get("sourceUrl"))
  row["_kind"] = "conflict"
  return row


def fillPolicyEvent(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  eventId = row["eventId"]
  blob = f"{eventId} {notes}"
  isWar = any(marker in blob for marker in WAR_NOTE_MARKERS)
  isReform = any(token in blob for token in ("改革", "reform"))
  isSakoku = "鎖国" in notes or "sakoku" in eventId
  if not row.get("promptForLeader"):
    row["promptForLeader"] = f"{notes}。法令と人心、財政への波及を見定めよ。"
  if not row.get("promptForOpinionLeader"):
    row["promptForOpinionLeader"] = f"{notes}、と市で噂になっている。"
  if isWar:
    row["scope"] = "japan"
    if row.get("disasterOverride") is None:
      row["disasterOverride"] = 0.7
    if row.get("populationShock") is None:
      row["populationShock"] = 0.006
    if row.get("cropLossExtra") is None:
      row["cropLossExtra"] = 0.03
    if row.get("socialUnrest") is None:
      row["socialUnrest"] = 0.12
  else:
    row["scope"] = "japan_info"
    if row.get("socialUnrest") is None and isReform:
      row["socialUnrest"] = 0.04
    if row.get("govDemand") is None and isReform:
      row["govDemand"] = 0.03
    if isSakoku and row.get("exportDrain") is None:
      row["exportDrain"] = 0.06
      row["fiatTrustShock"] = 0.02
  row["_sourceDecoded"] = decodeSource(row.get("sourceUrl"))
  row["_kind"] = "policy_event"
  return row


def parseLooseDate(raw: object) -> str | None:
  if raw is None or raw == "":
    return None
  if isinstance(raw, datetime):
    return f"{raw.year:04d}-{raw.month:02d}"
  text = str(raw).strip()
  if text in ("不明", "None", "nan"):
    return None
  match = re.search(r"(\d{4})(?:[-/年](\d{1,2}))?", text)
  if not match:
    return None
  year = int(match.group(1))
  month = int(match.group(2) or 1)
  return f"{year:04d}-{month:02d}"


def statuteTitle(notes: str, eventId: str) -> str:
  text = (notes or eventId).split("。")[0].strip()
  text = re.sub(r"公布:.*$", "", text).strip(" /")
  return text[:80] or eventId


def statuteEffects(notes: str) -> tuple[dict, dict]:
  effects: dict[str, float] = {"govDemand": 0.01}
  tweaks: dict = {"bureaucracyEfficiency": 1.02}
  if any(token in notes for token in ("税", "租税", "所得税", "関税")):
    effects["govDemand"] = 0.03
    tweaks["bureaucracyEfficiency"] = 1.06
  if any(token in notes for token in ("銀行", "通貨", "日本銀行", "紙幣", "金貨")):
    effects["fiatTrustShock"] = 0.03
  if any(token in notes for token in ("労働", "工場法")):
    effects["laborDrain"] = 0.01
    effects["socialUnrest"] = 0.02
  if any(token in notes for token in ("罰", "刑", "治罪", "爆発物", "警察")):
    effects["socialUnrest"] = 0.02
    tweaks["blackMarketCrackdown"] = 0.12
    tweaks["enforcementPriority"] = 1.1
  if any(token in notes for token in ("通商", "開港", "関税")):
    tweaks["tradeStance"] = "open"
    effects["importCostShock"] = 0.02
  if any(token in notes for token in ("配給", "統制", "食糧")):
    tweaks["foodRationPriority"] = "equal"
    tweaks["reserveReleaseRatio"] = 0.08
    effects["socialUnrest"] = 0.03
  if "廃止" in notes:
    effects["fiatTrustShock"] = effects.get("fiatTrustShock", 0.0) + 0.01
  return effects, tweaks


def fillStatute(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  promulgated = parseLooseDate(row.get("公布日"))
  enforced = parseLooseDate(row.get("施行日"))
  yearMonth = row["yearMonth"]
  availableFrom = promulgated or yearMonth
  historical = enforced or promulgated or yearMonth
  title = statuteTitle(notes, row["eventId"])
  effects, tweaks = statuteEffects(notes)
  source = decodeSource(row.get("sourceUrl"))
  row["title"] = title
  row["historicalYearMonth"] = historical
  row["availableFrom"] = availableFrom
  row["availableUntil"] = "2026-08"
  row["effects"] = effects
  row["policyTweaks"] = tweaks
  if not row.get("promptForLeader"):
    row["promptForLeader"] = f"{title}。公布・施行の時期に合わせて採用してもよい法令だ。"
  if not row.get("promptForOpinionLeader"):
    row["promptForOpinionLeader"] = f"{title}が届いた、と役人が噂する。"
  if source:
    row["notes"] = f"{notes} {source}".strip()
  row["_kind"] = "statute"
  return row


def sheetKind(sheetName: str) -> str:
  if "火山" in sheetName:
    return "volcano"
  if "疫病" in sheetName:
    return "epidemic"
  if "地震" in sheetName:
    return "earthquake"
  if "火事" in sheetName:
    return "fire"
  if "争い" in sheetName:
    return "conflict"
  if "政策" in sheetName:
    return "policy"
  return "infra"


def normalizeAndFill(rows: list[dict], kind: str) -> list[dict]:
  filled: list[dict] = []
  for row in rows:
    item = dict(row)
    try:
      item["yearMonth"] = normalizeYearMonth(item["yearMonth"])
    except ValueError:
      continue
    item["eventId"] = normalizeEventId(item["eventId"])
    item["scope"] = item.get("scope") or "japan"
    if item["scope"] not in ("japan", "japan_info", "world"):
      item["scope"] = "japan"
    if kind == "volcano":
      item = fillVolcano(item)
    elif kind == "epidemic":
      item = fillEpidemic(item)
    elif kind == "earthquake":
      item = fillEarthquake(item)
    elif kind == "fire":
      item = fillFire(item)
    elif kind == "conflict":
      item = fillConflict(item)
    elif kind == "policy":
      if item["eventId"].startswith("statute_"):
        item = fillStatute(item)
      else:
        item = fillPolicyEvent(item)
    else:
      item = fillInfra(item)
    filled.append(item)
  return filled


def preferredEventRow(left: dict, right: dict) -> dict:
  leftScore = (1 if left.get("disasterOverride") is not None else 0) + (
    1 if left.get("promptForLeader") else 0
  )
  rightScore = (1 if right.get("disasterOverride") is not None else 0) + (
    1 if right.get("promptForLeader") else 0
  )
  return left if leftScore >= rightScore else right


def collapseImported(rows: list[dict]) -> list[dict]:
  byId: dict[str, dict] = {}
  order: list[str] = []
  for row in rows:
    eventId = row["eventId"]
    if eventId not in byId:
      byId[eventId] = row
      order.append(eventId)
    else:
      byId[eventId] = preferredEventRow(byId[eventId], row)
  return [byId[eventId] for eventId in order]


def loadExistingCatalog() -> dict:
  path = JAPAN / "catalog.yaml"
  return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def loadExistingTimeline() -> list[dict]:
  path = JAPAN / "timeline.csv"
  with path.open(encoding="utf-8", newline="") as handle:
    return list(csv.DictReader(handle))


def writeCatalog(catalog: dict) -> None:
  text = yaml.safe_dump(
    catalog,
    allow_unicode=True,
    sort_keys=False,
    default_flow_style=False,
    width=1000,
  )
  (JAPAN / "catalog.yaml").write_text(text, encoding="utf-8")


def writeTimeline(rows: list[dict]) -> None:
  rowsSorted = sorted(rows, key=lambda row: (row["yearMonth"], row["eventId"]))
  with (JAPAN / "timeline.csv").open("w", encoding="utf-8", newline="") as handle:
    writer = csv.DictWriter(
      handle,
      fieldnames=["yearMonth", "eventId", "scope", "notes"],
      lineterminator="\n",
    )
    writer.writeheader()
    for row in rowsSorted:
      writer.writerow(
        {
          "yearMonth": row["yearMonth"],
          "eventId": row["eventId"],
          "scope": row.get("scope") or "japan",
          "notes": row.get("notes") or "",
        }
      )


def writeSources(rows: list[dict]) -> None:
  lines = [
    "Sources for shaped-sheet events (volcano / infrastructure / epidemic / earthquake / fire).",
    "Event blurbs paraphrase Japanese Wikipedia; see japan/README.md (CC BY-SA).",
    "",
  ]
  seen: set[str] = set()
  for row in rows:
    url = row.get("sourceUrl")
    if not url or url in seen:
      continue
    seen.add(url)
    lines.append(f"- {row['eventId']}: {decodeSource(url)}")
  (JAPAN / "sources_shaped_sheets.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


def catalogEntryFromRow(row: dict) -> dict:
  entry: dict = {
    "scope": row.get("scope") or "japan",
    "promptForLeader": row["promptForLeader"],
    "promptForOpinionLeader": row["promptForOpinionLeader"],
  }
  for key in (
    "disasterOverride",
    "populationShock",
    "cropLossExtra",
    "epidemicSeverity",
    "socialUnrest",
    "laborDrain",
    "infraDamage",
    "govDemand",
    "exportDrain",
    "fiatTrustShock",
  ):
    if row.get(key) is not None:
      entry[key] = float(row[key])
  if row.get("notes"):
    entry["notes"] = row["notes"]
  return entry


def policyEntryFromRow(row: dict) -> dict:
  entry: dict = {
    "title": row.get("title") or row["eventId"],
    "historicalYearMonth": row.get("historicalYearMonth") or row["yearMonth"],
    "availableFrom": row.get("availableFrom") or row["yearMonth"],
    "availableUntil": row.get("availableUntil") or "2026-08",
    "targetArea": "ALL",
    "promptForLeader": row["promptForLeader"],
    "promptForOpinionLeader": row["promptForOpinionLeader"],
    "effects": dict(row.get("effects") or {}),
    "policyTweaks": dict(row.get("policyTweaks") or {}),
  }
  if row.get("notes"):
    entry["notes"] = str(row["notes"])[:400]
  return entry


def loadExistingPolicies() -> dict:
  path = POLICIES / "catalog.yaml"
  if not path.exists():
    return {}
  return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def writePolicyCatalog(catalog: dict) -> None:
  POLICIES.mkdir(parents=True, exist_ok=True)
  text = yaml.safe_dump(
    catalog,
    allow_unicode=True,
    sort_keys=False,
    default_flow_style=False,
    width=1000,
  )
  (POLICIES / "catalog.yaml").write_text(text, encoding="utf-8")


def writeEarthquakePublicReadme(bibliography: list[str], quakeCount: int, fireCount: int) -> None:
  bibLines = "\n".join(f"- {url}" for url in bibliography) or "- (none)"
  text = f"""# 日本イベント（公開してよい出典メモ）

このフォルダの `catalog.yaml` / `timeline.csv` はシミュレーション用の短いプロンプトと係数です。
Wikipedia記事本文の全文転載ではありません。

## 地震シートの参考文献（一覧サイト）

地震の**件名が存在するかどうか**の確認に、次のページを見た。各地震の内容は日本語版 Wikipedia の当該記事から要約している。

{bibLines}

（[Nippon.com の当該特集](https://www.nippon.com/ja/features/h20002/) は著作物なので、記事本文はコピーしていない。URL を参考文献として示すだけ。）

## Wikipedia（CC BY-SA）

各行の `sourceUrl` は主に `https://ja.wikipedia.org/wiki/...`。
公開リポジトリに載せるのは:

- こちらで書き直した短い日本語プロンプト（オリジナル要約）
- 記事へのリンク一覧（`sources_shaped_sheets.txt`）

これなら **CC BY-SA の帰属（リンク）** として問題になりにくい。
Wikipedia本文を長文のまま YAML に貼るのは避ける（シェアAlikeと分量の問題）。

件数: 地震 {quakeCount} 件、火事 {fireCount} 件（取り込み時点）。
"""
  (JAPAN / "README.md").write_text(text, encoding="utf-8")


def parseArgs() -> argparse.Namespace:
  parser = argparse.ArgumentParser(
    description="Import shaped / conflict / policy sheets from a local xlsx."
  )
  parser.add_argument(
    "--xlsx",
    default=None,
    help=f"Workbook path (else env {ENV_XLSX}, else default search paths; see script source)",
  )
  return parser.parse_args()


def main() -> int:
  args = parseArgs()
  xlsxPath = resolveXlsx(args.xlsx)

  LOGS.mkdir(exist_ok=True)
  backupDir = LOGS / "backup_japan_events_before_conflict_policy"
  backupDir.mkdir(exist_ok=True)
  for name in ("catalog.yaml", "timeline.csv"):
    src = JAPAN / name
    if src.exists():
      shutil.copy2(src, backupDir / name)
  policySrc = POLICIES / "catalog.yaml"
  if policySrc.exists():
    shutil.copy2(policySrc, backupDir / "policies_catalog.yaml")

  sheets, bibliography = dumpShapedSheets(xlsxPath)
  imported: list[dict] = []
  statutes: list[dict] = []
  counts: dict[str, int] = {}
  for sheetName, raw in sheets.items():
    kind = sheetKind(sheetName)
    filled = normalizeAndFill(raw, kind)
    counts[sheetName] = len(filled)
    for row in filled:
      if row.get("_kind") == "statute":
        statutes.append(row)
      else:
        imported.append(row)
  imported = collapseImported(imported)
  statutes = collapseImported(statutes)

  (LOGS / "import_shaped_sheets_filled.json").write_text(
    json.dumps(imported[:80], ensure_ascii=False, indent=2, default=str),
    encoding="utf-8",
  )

  catalog = loadExistingCatalog()
  for row in imported:
    catalog[row["eventId"]] = catalogEntryFromRow(row)

  timeline = loadExistingTimeline()
  existingKeys = {(row["yearMonth"], row["eventId"]) for row in timeline}
  for row in imported:
    if row["yearMonth"] < SIM_START:
      continue
    key = (row["yearMonth"], row["eventId"])
    if key in existingKeys:
      continue
    timeline.append(
      {
        "yearMonth": row["yearMonth"],
        "eventId": row["eventId"],
        "scope": row.get("scope") or "japan",
        "notes": row.get("notes") or "",
      }
    )
    existingKeys.add(key)

  policyCatalog = loadExistingPolicies()
  for row in statutes:
    policyId = row["eventId"]
    if policyId in policyCatalog:
      continue
    policyCatalog[policyId] = policyEntryFromRow(row)

  writeCatalog(catalog)
  writeTimeline(timeline)
  writePolicyCatalog(policyCatalog)
  writeSources(imported)
  quakeCount = sum(1 for row in imported if row.get("_kind") == "earthquake")
  fireCount = sum(1 for row in imported if row.get("_kind") == "fire")
  writeEarthquakePublicReadme(bibliography, quakeCount, fireCount)

  print(f"xlsx={xlsxPath}")
  print("imported " + " ".join(f"{name}={n}" for name, n in counts.items()))
  print(
    f"japan_new={len(imported)} statutes={len(statutes)} "
    f"catalog={len(catalog)} timeline={len(timeline)} policies={len(policyCatalog)} backup={backupDir}"
  )
  print("bibliography=" + ", ".join(bibliography))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
