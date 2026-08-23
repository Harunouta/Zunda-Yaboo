"""Import 飢饉 + 水害 from categorized_events.xlsx into japan events.

Fills blank prompts / coefficients from notes + Wikipedia-informed heuristics.
Does not import 政治 / 地震 / 火山 sheets yet — extend SHEETS (or add
import_categorized_sheets.py) when the user delivers those lists. See HANDOFF.md.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
from pathlib import Path
from urllib.parse import unquote

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ENV_XLSX = "ZUNDA_EVENTS_XLSX"
ENV_SPEC = "ZUNDA_EVENTS_SPEC"
DEFAULT_XLSX_NAME = "categorized_events.xlsx"
DEFAULT_SPEC_NAME = "sikii.md"
JAPAN = ROOT / "data" / "events" / "japan"
LOGS = ROOT / "logs"

SHEETS = ("水害", "その他の農作物への被害")
REPLACE_TIMELINE_IDS = frozenset({"tenmei_famine", "tenpo_famine"})

SEVERE_FLOOD_MARKERS = (
  "isewan",
  "kathleen",
  "makurazaki",
  "muroto",
  "hanshin",
  "tokyo-fullwater",
  "nagasaki",
  "westjapan",
  "h30west",
  "r1_typhoon",
  "dog_fullwater",
  "yokota",
  "kanogawa",
  "isahaya",
  "kishu",
  "tooyamaru",
  "oki_typhoon",
)

def normalizeYearMonth(raw: str) -> str:
  text = str(raw).strip().replace("_", "-")
  yearText, monthText = text.split("-", 1)
  return f"{int(yearText):04d}-{int(monthText):02d}"


def normalizeEventId(raw: str) -> str:
  text = str(raw).strip()
  text = text.replace("-", "_").replace(" ", "_")
  text = re.sub(r"[^0-9A-Za-z_]", "_", text)
  text = re.sub(r"_+", "_", text).strip("_")
  return text.lower()


def decodeSource(url: str | None) -> str:
  if not url:
    return ""
  return unquote(url)


def resolveXlsx(cliPath: str | None) -> Path:
  if cliPath:
    path = Path(cliPath)
    if not path.is_file():
      raise SystemExit(f"missing xlsx: {path}")
    return path
  envPath = os.getenv(ENV_XLSX)
  if envPath:
    path = Path(envPath)
    if not path.is_file():
      raise SystemExit(f"missing xlsx ({ENV_XLSX}): {path}")
    return path
  path = Path.home() / "Downloads" / DEFAULT_XLSX_NAME
  if path.is_file():
    return path
  raise SystemExit(f"missing xlsx: pass --xlsx PATH or set {ENV_XLSX} (tried {path})")


def resolveSpec(cliPath: str | None) -> Path | None:
  if cliPath:
    path = Path(cliPath)
    if not path.is_file():
      raise SystemExit(f"missing spec: {path}")
    return path
  envPath = os.getenv(ENV_SPEC)
  if envPath:
    path = Path(envPath)
    if not path.is_file():
      raise SystemExit(f"missing spec ({ENV_SPEC}): {path}")
    return path
  path = Path.home() / "Downloads" / DEFAULT_SPEC_NAME
  if path.is_file():
    return path
  return None


def dumpSheet(xlsxPath: Path, sheetName: str) -> list[dict]:
  workbook = load_workbook(xlsxPath, read_only=True, data_only=True)
  worksheet = workbook[sheetName]
  rows = list(worksheet.iter_rows(values_only=True))
  header = [str(cell) for cell in rows[0]]
  out: list[dict] = []
  for row in rows[1:]:
    if not row or row[0] is None or str(row[0]).strip() == "":
      continue
    entry: dict = {}
    for key, value in zip(header, row):
      if value is None or value == "":
        entry[key] = None
      elif key in ("disasterOverride", "populationShock", "cropLossExtra"):
        entry[key] = float(value)
      else:
        entry[key] = str(value).strip()
    if entry.get("yearMonth") and entry.get("eventId"):
      out.append(entry)
  return out


def seriesPhase(eventId: str) -> str:
  match = re.search(r"_(\d+)$", eventId)
  if not match:
    return "single"
  index = int(match.group(1))
  if index <= 2:
    return "peak"
  if index <= 5:
    return "severe"
  return "linger"


def fillFamine(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  eventId = row["eventId"]
  phase = seriesPhase(eventId)
  source = decodeSource(row.get("sourceUrl"))

  if "寛永" in notes or eventId.startswith("kanei"):
    leader = "異常気象による冷害・干ばつ・虫害が重なり大飢饉となった。諸大名に救済を命じよ。"
    opinion = "雨が降らず麦が枯れ、雨が降れば虫が湧く。もう生きていけない。"
    coeffs = (0.45, 0.01, 0.15)
  elif "享保" in notes or eventId.startswith("kyoho"):
    leader = "西国でウンカの大発生により未曾有の凶作となった。備蓄米を放出し餓死者を防げ。"
    opinion = "虫が稲を食い尽くした。西国から逃げてきた人が増えている。"
    coeffs = (0.4, 0.01, 0.15)
  elif "天明" in notes or "tenmei" in eventId:
    leader = {
      "peak": "天明の大飢饉の只中だ。浅間の灰と冷害で米が枯れる。施米と流通を急げ。",
      "severe": "天明飢饉が続く。餓死者と一揆の兆しがある。備蓄と治安を両立せよ。",
      "linger": "天明の余波で農村が弱っている。回復と備蓄再建を急げ。",
      "single": "天明の大飢饉だ。備蓄を配り、米価を抑えよ。",
    }[phase]
    opinion = {
      "peak": "空が暗い。隣村で倒れる人がいる。米が手に入らない。",
      "severe": "飢えた行商人が増えた。闇市か逃げるかしかない。",
      "linger": "まだ腹が減る。凶作の噂が止まらない。",
      "single": "空が暗い。通貨が腐る呪いが噂されている。",
    }[phase]
    coeffs = {
      "peak": (0.35, 0.015, 0.2),
      "severe": (0.42, 0.01, 0.14),
      "linger": (0.55, 0.005, 0.08),
      "single": (0.35, 0.015, 0.2),
    }[phase]
  elif "天保" in notes or "tenpo" in eventId:
    leader = {
      "peak": "天保の大飢饉。大雨と冷害で全国的に大凶作だ。江戸への米を確保し打ちこわしを防げ。",
      "severe": "天保飢饉が続く。法令と施しで民を繋ぎ止めよ。",
      "linger": "天保の凶作余波。米価と治安を安定させよ。",
      "single": "天保の飢饉。民を救う法令を出すのだ。",
    }[phase]
    opinion = {
      "peak": "米屋の前に人が群がる。明日の飯がない。",
      "severe": "飢えた行商人として、闇市か逃げるか決めろ。",
      "linger": "まだ高い。凶作が終わったとは思えない。",
      "single": "飢えた行商人として、闇市か逃げるか決めろ。",
    }[phase]
    coeffs = {
      "peak": (0.38, 0.012, 0.18),
      "severe": (0.45, 0.008, 0.12),
      "linger": (0.55, 0.004, 0.07),
      "single": (0.38, 0.012, 0.18),
    }[phase]
  elif "宝暦" in notes or "horeki" in eventId:
    leader = {
      "peak": "宝暦の飢饉。冷害と洪水で東北・関東の作柄が崩れた。救済と年貢減免を検討せよ。",
      "severe": "宝暦の凶作が続く。地方からの報を集め、米を回せ。",
      "linger": "宝暦飢饉の余波。農村の立て直しを急げ。",
      "single": "宝暦の飢饉。救済を命じよ。",
    }[phase]
    opinion = "冷えて実が入らない。種籾まで食う話が出ている。"
    coeffs = {
      "peak": (0.48, 0.007, 0.12),
      "severe": (0.55, 0.004, 0.08),
      "linger": (0.65, 0.002, 0.05),
      "single": (0.5, 0.006, 0.1),
    }[phase]
  elif "元禄" in notes or "genroku" in eventId:
    # Wikipedia 東北凶作: multi-year cold damage / crop failure in the Genroku era window.
    leader = {
      "peak": "東北を中心に凶作が広がっている（元禄期の冷害）。備蓄放出と移入米を急げ。",
      "severe": "元禄期の東北凶作が続く。年貢と救済のバランスを取れ。",
      "linger": "凶作の余波で農村が疲弊している。回復策を打て。",
      "single": "東北の凶作。民を飢えさせるな。",
    }[phase]
    opinion = "北の村から飢えた人が来る。米が無いと噂だ。"
    coeffs = {
      "peak": (0.5, 0.006, 0.1),
      "severe": (0.58, 0.003, 0.07),
      "linger": (0.68, 0.002, 0.04),
      "single": (0.55, 0.004, 0.08),
    }[phase]
  elif "1905" in notes or eventId.startswith("1905"):
    leader = "1905年の大凶作。冷害で東北の米が壊滅した。移入と価格安定を急げ。"
    opinion = "東北は作が無い。米が暴騰すると噂だ。"
    coeffs = (0.55, 0.004, 0.12)
  elif "豊作飢饉" in notes or eventId.startswith("full_famine"):
    leader = "豊作なのに価格崩壊で農家が潰れる（昭和初期の農業恐慌前兆）。価格と流通を整えよ。"
    opinion = "米は余っているのに金にならない。百姓が立ち行かない。"
    coeffs = (0.75, 0.001, 0.02)
  elif "昭和農業恐慌" in notes or "showa_famine" in eventId:
    leader = "昭和農業恐慌。デフレと米価低迷で農村が破綻寸前だ。救済融資と需給調整を。"
    opinion = "作っても赤字。娘の身売り話が村を走る。"
    coeffs = (0.7, 0.002, 0.03)
  elif "1934" in notes or "1934_famine" in eventId:
    leader = "1934年の大凶作。冷害で東北が再び壊滅した。緊急移入と救農を急げ。"
    opinion = "また東北が飢える。米が高くて手が出せない。"
    coeffs = (0.52, 0.005, 0.14) if phase in ("peak", "single") else (0.6, 0.003, 0.08)
  elif "平成の米" in notes or "heisei_rice" in eventId:
    leader = "記録的冷夏で米が歴史的凶作となった。緊急輸入を決め、市場パニックを抑えよ。"
    opinion = "店に米が無い。外国米の噂が飛び交う。"
    coeffs = (0.7, None, 0.3)
  else:
    leader = f"{notes}。凶作・飢饉への救済と米の流通確保を急げ。"
    opinion = "腹が減る。作が悪いと噂が広がる。"
    coeffs = (0.55, 0.004, 0.08)

  if not row.get("promptForLeader"):
    row["promptForLeader"] = leader
  if not row.get("promptForOpinionLeader"):
    row["promptForOpinionLeader"] = opinion
  if row.get("disasterOverride") is None:
    row["disasterOverride"] = coeffs[0]
  if row.get("populationShock") is None and coeffs[1] is not None:
    row["populationShock"] = coeffs[1]
  if row.get("cropLossExtra") is None and coeffs[2] is not None:
    row["cropLossExtra"] = coeffs[2]
  row["_sourceDecoded"] = source
  return row


def isSevereFlood(eventId: str, notes: str) -> bool:
  blob = f"{eventId} {notes}".lower()
  return any(marker in blob for marker in SEVERE_FLOOD_MARKERS)


def fillFlood(row: dict) -> dict:
  notes = row.get("notes") or row["eventId"]
  eventId = row["eventId"]
  source = decodeSource(row.get("sourceUrl"))
  severe = isSevereFlood(eventId, notes)
  isSnow = "雪" in notes or "snow" in eventId.lower()
  isTyphoon = "台風" in notes or "typhoon" in eventId.lower()

  if eventId.startswith("dog_fullwater") or "寛保" in notes:
    leader = "寛保の洪水・高潮。利根・荒川流域が広範囲に浸水した。救恤と堤防補修を急げ。"
    opinion = "水が堤を越え、村が沈んだ。舟でしか動けない。"
    coeffs = (0.42, 0.006, 0.12)
  elif "淀川" in notes or eventId.lower().startswith(("kyodo", "myodo")):
    leader = f"{notes}。淀川が氾濫し畿内の田畑が水没した。排水と救済を命じよ。"
    opinion = "川が暴れた。田も家も泥だらけだ。"
    coeffs = (0.5, 0.003, 0.1)
  elif "横田" in notes or "yokota" in eventId.lower():
    leader = "利根川の横田切れ。関東の低地が広範囲に浸水した。救援と復旧を急げ。"
    opinion = "堤が切れた。水がどこまで来るかわからない。"
    coeffs = (0.48, 0.004, 0.11)
  elif "明治40" in notes or "myamanashi" in eventId.lower():
    leader = f"{notes}（富士川・甲府盆地周辺の水害系列）。土砂と洪水への対応を急げ。"
    opinion = "山が崩れ、川が黒い。逃げ遅れた人がいる。"
    coeffs = (0.55, 0.002, 0.08)
  elif "伊勢湾" in notes or "isewan" in eventId.lower():
    leader = "伊勢湾台風。高潮と暴風で甚大な被害が出た。国家規模の救援を命じよ。"
    opinion = "海が街を飲み込んだ。夜の闇の中で水が来た。"
    coeffs = (0.5, 0.008, 0.1)
  elif "カスリーン" in notes or "kathleen" in eventId.lower():
    leader = "カスリーン台風。利根川が決壊し関東平野が水没した。救援と堤防再建を急げ。"
    opinion = "関東一帯が湖になった。屋根の上で救助を待つ。"
    coeffs = (0.45, 0.007, 0.12)
  elif "枕崎" in notes or "makurazaki" in eventId.lower():
    leader = "枕崎台風。終戦直後の西日本を暴風・豪雨が襲った。救援物資の配分を急げ。"
    opinion = "戦争の後に台風。家も食料も足りない。"
    coeffs = (0.48, 0.006, 0.1)
  elif "室戸" in notes or "muroto" in eventId.lower():
    leader = f"{notes}。暴風と高潮で西日本が壊滅級の被害を受けた。救護と復旧を。"
    opinion = "風が家を壊し、潮が街を洗う。生きた心地がしない。"
    coeffs = (0.48, 0.006, 0.1)
  elif isSnow:
    leader = f"{notes}。交通・物流と除雪、孤立集落の救援を優先せよ。"
    opinion = "雪が止まない。道が消え、食料が届かない。"
    coeffs = (0.72, 0.001, 0.03)
  elif isTyphoon:
    leader = f"{notes}。暴風・大雨・高潮への警戒と救援を命じよ。"
    opinion = "台風だ。屋根が飛び、川が溢れる。"
    coeffs = (0.5, 0.004, 0.09) if severe else (0.58, 0.002, 0.06)
  elif severe:
    leader = f"{notes}。広域水害だ。堤防・排水・救恤を国家規模で進めよ。"
    opinion = "水が来た。田も家も沈む。逃げるしかない。"
    coeffs = (0.48, 0.005, 0.1)
  else:
    leader = f"{notes}。浸水被害への救援と農地復旧を急げ。"
    opinion = "川が増水した。床上まで水が来ると噂だ。"
    coeffs = (0.6, 0.002, 0.06)

  if not row.get("promptForLeader"):
    row["promptForLeader"] = leader
  if not row.get("promptForOpinionLeader"):
    row["promptForOpinionLeader"] = opinion
  if row.get("disasterOverride") is None:
    row["disasterOverride"] = coeffs[0]
  if row.get("populationShock") is None:
    row["populationShock"] = coeffs[1]
  if row.get("cropLossExtra") is None:
    row["cropLossExtra"] = coeffs[2]
  row["_sourceDecoded"] = source
  return row


def normalizeAndFill(rows: list[dict], kind: str) -> list[dict]:
  filled: list[dict] = []
  seenIds: set[str] = set()
  for row in rows:
    item = dict(row)
    item["yearMonth"] = normalizeYearMonth(item["yearMonth"])
    item["eventId"] = normalizeEventId(item["eventId"])
    if item["eventId"] in seenIds:
      # Same catalog id fired in multiple months is OK; keep one catalog body.
      pass
    seenIds.add(item["eventId"])
    item["scope"] = item.get("scope") or "japan"
    if kind == "famine":
      item = fillFamine(item)
    else:
      item = fillFlood(item)
    filled.append(item)
  return filled


def loadExistingCatalog() -> dict:
  path = JAPAN / "catalog.yaml"
  return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def loadExistingTimeline() -> list[dict]:
  path = JAPAN / "timeline.csv"
  with path.open(encoding="utf-8", newline="") as handle:
    return list(csv.DictReader(handle))


def writeCatalog(catalog: dict) -> None:
  # Prefer block style for readability of Japanese prompts.
  text = yaml.safe_dump(
    catalog,
    allow_unicode=True,
    sort_keys=False,
    default_flow_style=False,
    width=1000,
  )
  (JAPAN / "catalog.yaml").write_text(text, encoding="utf-8")


def writeTimeline(rows: list[dict]) -> None:
  rowsSorted = sorted(rows, key=lambda row: (row["yearMonth"], row["eventId"]))
  with (JAPAN / "timeline.csv").open("w", encoding="utf-8", newline="") as handle:
    writer = csv.DictWriter(
      handle,
      fieldnames=["yearMonth", "eventId", "scope", "notes"],
      lineterminator="\n",
    )
    writer.writeheader()
    for row in rowsSorted:
      writer.writerow(
        {
          "yearMonth": row["yearMonth"],
          "eventId": row["eventId"],
          "scope": row.get("scope") or "japan",
          "notes": row.get("notes") or "",
        }
      )


def writeSources(rows: list[dict]) -> None:
  lines = [
    "Sources for famine/flood events imported into Zunda-Yaboo.",
    "Many entries summarize or paraphrase Japanese Wikipedia (CC BY-SA).",
    "See each URL for attribution of the underlying article.",
    "",
  ]
  seen: set[str] = set()
  for row in rows:
    url = row.get("sourceUrl")
    if not url or url in seen:
      continue
    seen.add(url)
    lines.append(f"- {row['eventId']}: {decodeSource(url)}")
  (JAPAN / "sources_famine_flood.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")


def catalogEntryFromRow(row: dict) -> dict:
  entry: dict = {
    "scope": row.get("scope") or "japan",
    "promptForLeader": row["promptForLeader"],
    "promptForOpinionLeader": row["promptForOpinionLeader"],
  }
  if row.get("disasterOverride") is not None:
    entry["disasterOverride"] = float(row["disasterOverride"])
  if row.get("populationShock") is not None:
    entry["populationShock"] = float(row["populationShock"])
  if row.get("cropLossExtra") is not None:
    entry["cropLossExtra"] = float(row["cropLossExtra"])
  if row.get("notes"):
    entry["notes"] = row["notes"]
  return entry


def parseArgs() -> argparse.Namespace:
  parser = argparse.ArgumentParser(description="Import famine and flood sheets from a local xlsx.")
  parser.add_argument(
    "--xlsx",
    default=None,
    help=f"Workbook path (else {ENV_XLSX}, else ~/Downloads/{DEFAULT_XLSX_NAME})",
  )
  parser.add_argument(
    "--spec",
    default=None,
    help=f"Optional sikii.md to copy into UNIFIED_EVENTS_SPEC.md (else {ENV_SPEC})",
  )
  return parser.parse_args()


def main() -> int:
  args = parseArgs()
  xlsxPath = resolveXlsx(args.xlsx)
  specPath = resolveSpec(args.spec)

  LOGS.mkdir(exist_ok=True)
  backupDir = LOGS / "backup_japan_events_before_famine_flood"
  backupDir.mkdir(exist_ok=True)
  for name in ("catalog.yaml", "timeline.csv"):
    src = JAPAN / name
    if src.exists():
      shutil.copy2(src, backupDir / name)

  if specPath is not None:
    shutil.copy2(specPath, ROOT / "data" / "events" / "UNIFIED_EVENTS_SPEC.md")

  floodRaw = dumpSheet(xlsxPath, "水害")
  famineRaw = dumpSheet(xlsxPath, "その他の農作物への被害")
  flood = normalizeAndFill(floodRaw, "flood")
  famine = normalizeAndFill(famineRaw, "famine")
  # Keep a late Tenmei linger month that the sparse timeline used to fire.
  lateTenmei = fillFamine(
    {
      "yearMonth": "1787-05",
      "eventId": "tenmei_famine_6",
      "scope": "japan",
      "notes": "天明の大飢饉_余波",
      "promptForLeader": None,
      "promptForOpinionLeader": None,
      "disasterOverride": None,
      "populationShock": None,
      "cropLossExtra": None,
      "sourceUrl": "https://ja.wikipedia.org/wiki/天明の大飢饉",
    }
  )
  lateTenmei["yearMonth"] = normalizeYearMonth(lateTenmei["yearMonth"])
  lateTenmei["eventId"] = normalizeEventId(lateTenmei["eventId"])
  famine.append(lateTenmei)
  imported = famine + flood

  (LOGS / "import_famine_flood_filled.json").write_text(
    json.dumps(imported, ensure_ascii=False, indent=2),
    encoding="utf-8",
  )

  catalog = loadExistingCatalog()
  # Drop sparse famine defs that the numbered series replaces.
  for staleId in REPLACE_TIMELINE_IDS:
    catalog.pop(staleId, None)

  for row in imported:
    catalog[row["eventId"]] = catalogEntryFromRow(row)

  # Keep alias stubs so old docs / smoke references stay readable.
  if "tenmei_famine_1" in catalog:
    catalog["tenmei_famine"] = dict(catalog["tenmei_famine_1"])
    catalog["tenmei_famine"]["notes"] = "alias of tenmei_famine_1"
  if "tenpo_famine_1" in catalog:
    catalog["tenpo_famine"] = dict(catalog["tenpo_famine_1"])
    catalog["tenpo_famine"]["notes"] = "alias of tenpo_famine_1"

  timeline = loadExistingTimeline()
  kept = [
    row
    for row in timeline
    if row.get("eventId") not in REPLACE_TIMELINE_IDS
  ]
  existingKeys = {(row["yearMonth"], row["eventId"]) for row in kept}
  for row in imported:
    key = (row["yearMonth"], row["eventId"])
    if key in existingKeys:
      continue
    kept.append(
      {
        "yearMonth": row["yearMonth"],
        "eventId": row["eventId"],
        "scope": row.get("scope") or "japan",
        "notes": row.get("notes") or "",
      }
    )
    existingKeys.add(key)

  writeCatalog(catalog)
  writeTimeline(kept)
  writeSources(imported)

  print(
    f"imported famine={len(famine)} flood={len(flood)} "
    f"catalog={len(catalog)} timeline={len(kept)}"
  )
  print(f"backup={backupDir}")
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
