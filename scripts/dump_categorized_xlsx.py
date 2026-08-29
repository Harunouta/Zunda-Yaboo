"""Dump flood/famine sheets from categorized_events.xlsx for import."""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path

from openpyxl import load_workbook

ENV_XLSX = "ZUNDA_EVENTS_XLSX"
DEFAULT_XLSX_NAME = "categorized_events.xlsx"
OUT = Path(__file__).resolve().parents[1] / "logs"


def resolveXlsx(cliPath: str | None) -> Path:
  if cliPath:
    path = Path(cliPath)
    if not path.is_file():
      raise SystemExit(f"missing xlsx: {path}")
    return path
  envPath = os.getenv(ENV_XLSX)
  if envPath:
    path = Path(envPath)
    if not path.is_file():
      raise SystemExit(f"missing xlsx ({ENV_XLSX}): {path}")
    return path
  path = Path.home() / "Downloads" / DEFAULT_XLSX_NAME
  if path.is_file():
    return path
  raise SystemExit(f"missing xlsx: pass --xlsx PATH or set {ENV_XLSX} (tried {path})")


def dumpSheet(xlsxPath: Path, sheetName: str) -> list[dict]:
  workbook = load_workbook(xlsxPath, read_only=True, data_only=True)
  worksheet = workbook[sheetName]
  rows = list(worksheet.iter_rows(values_only=True))
  header = [str(cell) for cell in rows[0]]
  out: list[dict] = []
  for row in rows[1:]:
    if not row or row[0] is None or str(row[0]).strip() == "":
      continue
    entry: dict = {}
    for key, value in zip(header, row):
      if value is None or value == "":
        entry[key] = None
      elif key in ("disasterOverride", "populationShock", "cropLossExtra"):
        entry[key] = float(value)
      else:
        entry[key] = str(value).strip()
    if entry.get("yearMonth") and entry.get("eventId"):
      out.append(entry)
  return out


def parseArgs() -> argparse.Namespace:
  parser = argparse.ArgumentParser(description="Dump flood/famine xlsx sheets to logs JSON.")
  parser.add_argument(
    "--xlsx",
    default=None,
    help=f"Workbook path (else env {ENV_XLSX}, else default search paths; see script source)",
  )
  return parser.parse_args()


def main() -> None:
  xlsxPath = resolveXlsx(parseArgs().xlsx)
  OUT.mkdir(exist_ok=True)
  for sheetName, fileName in [
    ("水害", "flood"),
    ("その他の農作物への被害", "famine"),
  ]:
    rows = dumpSheet(xlsxPath, sheetName)
    filled = sum(1 for row in rows if row.get("promptForLeader"))
    print(sheetName, "total", len(rows), "filled", filled, "empty", len(rows) - filled)
    (OUT / f"import_{fileName}_raw.json").write_text(
      json.dumps(rows, ensure_ascii=False, indent=2),
      encoding="utf-8",
    )
    for row in rows:
      if not row.get("promptForLeader"):
        print(
          "EMPTY",
          row["yearMonth"],
          row["eventId"],
          row.get("notes"),
          row.get("sourceUrl"),
        )


if __name__ == "__main__":
  main()
